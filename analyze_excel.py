import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('DATA.XLSX')
sheet = wb.active

# Get headers (first row)
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print("=" * 60)
print("EXCEL DOSYASI ANALİZİ - DATA.XLSX")
print("=" * 60)
print(f"\n📋 Toplam Kolon Sayısı: {len(headers)}")
print(f"📊 Toplam Satır Sayısı: {sheet.max_row - 1}")  # Excluding header

print("\n" + "=" * 60)
print("KOLONLAR:")
print("=" * 60)
for i, h in enumerate(headers, 1):
    print(f"{i:2d}. {h}")

# Sample first 3 data rows
print("\n" + "=" * 60)
print("ÖRNEK VERİ (İlk 3 Satır):")
print("=" * 60)

for row_idx in range(2, min(5, sheet.max_row + 1)):  # Rows 2-4 (first 3 data rows)
    print(f"\n--- SATIR {row_idx - 1} ---")
    for col_idx, header in enumerate(headers, 1):
        cell_value = sheet.cell(row_idx, col_idx).value
        print(f"  {header}: {cell_value}")

print("\n" + "=" * 60)
