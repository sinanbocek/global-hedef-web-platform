import openpyxl
from collections import Counter

# Load the Excel file
wb = openpyxl.load_workbook('DATA.XLSX')
sheet = wb.active

# Get headers
headers = [cell.value for cell in sheet[1]]

print("=" * 80)
print("VERİ KALİTE ANALİZİ - DATA.XLSX")
print("=" * 80)

# Collect all unique values for key columns
salespeople = set()
companies = set()
policy_types = set()
customer_types = set()
customers = set()
tckn_lengths = Counter()

for row_idx in range(2, sheet.max_row + 1):
    # Satışçılar
    salesperson = sheet.cell(row_idx, 11).value  # SATIŞÇI
    if salesperson:
        salespeople.add(str(salesperson).strip())
    
    # Şirketler
    company = sheet.cell(row_idx, 13).value  # ŞİRKET
    if company:
        companies.add(str(company).strip())
    
    # Poliçe Türleri
    policy_type = sheet.cell(row_idx, 5).value  # POLİÇE TÜRÜ
    if policy_type:
        policy_types.add(str(policy_type).strip())
    
    # Müşteri Türü
    cust_type = sheet.cell(row_idx, 1).value  # Müşteri Türü
    if cust_type:
        customer_types.add(str(cust_type).strip())
    
    # Müşteriler
    customer = sheet.cell(row_idx, 2).value  # MÜŞTERİ
    if customer:
        customers.add(str(customer).strip())
    
    # TCKN uzunlukları
    tckn = sheet.cell(row_idx, 3).value  # TCKN
    if tckn:
        tckn_str = str(tckn).strip()
        tckn_lengths[len(tckn_str)] += 1

print(f"\n📊 Toplam Veri Sayısı: {sheet.max_row - 1} satır")
print(f"📊 Tekil Müşteri: {len(customers)} kişi/kurum")

print("\n" + "=" * 80)
print("👥 SATIŞÇILAR ({} kişi)".format(len(salespeople)))
print("=" * 80)
for i, sp in enumerate(sorted(salespeople), 1):
    print(f"{i:2d}. {sp}")

print("\n" + "=" * 80)
print("🏢 SİGORTA ŞİRKETLERİ ({} şirket)".format(len(companies)))
print("=" * 80)
for i, comp in enumerate(sorted(companies), 1):
    print(f"{i:2d}. {comp}")

print("\n" + "=" * 80)
print("📋 POLİÇE TÜRLERİ ({} tür)".format(len(policy_types)))
print("=" * 80)
for i, pt in enumerate(sorted(policy_types), 1):
    print(f"{i:2d}. {pt}")

print("\n" + "=" * 80)
print("👤 MÜŞTERİ TÜRLERİ")
print("=" * 80)
for ct in sorted(customer_types):
    print(f"  • {ct}")

print("\n" + "=" * 80)
print("🔢 TCKN/VKN UZUNLUK DAĞILIMI")
print("=" * 80)
for length in sorted(tckn_lengths.keys()):
    count = tckn_lengths[length]
    percentage = (count / (sheet.max_row - 1)) * 100
    print(f"  {length} hane: {count:3d} kayıt ({percentage:5.1f}%)")

print("\n" + "=" * 80)
print("📝 ÖRNEK MÜŞTERİ İSİMLERİ (İlk 10)")
print("=" * 80)
for i, cust in enumerate(sorted(customers)[:10], 1):
    print(f"{i:2d}. {cust}")

print("\n" + "=" * 80)
print("⚠️ VERİ KALİTE KONTROL")
print("=" * 80)

# Check for missing critical data
missing_policy_no = 0
missing_dates = 0
missing_customer = 0
null_premium = 0
null_commission = 0
potansiyel_count = 0

for row_idx in range(2, sheet.max_row + 1):
    if not sheet.cell(row_idx, 2).value:  # MÜŞTERİ
        missing_customer += 1
    
    policy_no = sheet.cell(row_idx, 6).value  # POLİÇE NUMARASI
    if not policy_no or str(policy_no).strip().upper() == 'POTANSİYEL':
        potansiyel_count += 1
    
    if not sheet.cell(row_idx, 7).value or not sheet.cell(row_idx, 8).value:  # Tarihler
        missing_dates += 1
    
    if sheet.cell(row_idx, 10).value is None:  # PRİM
        null_premium += 1
    
    if sheet.cell(row_idx, 12).value is None:  # KOMİSYON
        null_commission += 1

print(f"  ❌ Müşteri Adı Eksik: {missing_customer}")
print(f"  ⚠️  Potansiyel Poliçe: {potansiyel_count}")
print(f"  ❌ Tarih Eksik: {missing_dates}")
print(f"  ⚠️  Prim NULL: {null_premium}")
print(f"  ⚠️  Komisyon NULL: {null_commission}")

print("\n" + "=" * 80)
